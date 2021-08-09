from openpyxl import Workbook, load_workbook
import os

west_dir = "C:\\Users\\Carroll Boone Water\\Desktop\\2021_ADH_Reports\\West_Plant_Operations_Reports\\"
east_dir = "C:\\Users\\Carroll Boone Water\\Desktop\\2021_ADH_Reports\\East_Plant_Operations_Reports\\"
west_aug = "08-2021_CBWD_West_FR.xlsx"
east_aug = "08-2021_CBWD_East_FR.xlsx"
west_sept = "09-2021_CBWD_West_FR.xlsx"
east_sept = "09-2021_CBWD_East_FR.xlsx"
west_aug_path = west_dir + west_aug
east_aug_path = east_dir + east_aug
west_sept_path = west_dir + west_sept
east_sept_path = east_dir + east_sept


if __name__ == '__main__':
	print(west_aug_path)

	west_wb = load_workbook(west_aug_path)
	east_wb = load_workbook(east_aug_path)

	w_active = west_wb.active
	e_active = east_wb.active

	print(w_active['G9'].value)
	print(w_active['G10'].value)

	G9 = w_active['G9'].value
	G10 = w_active['G10'].value
	G11 = w_active['G11'].value
	G12 = w_active['G12'].value
	G13 = w_active['G13'].value
	G14 = w_active['G14'].value
	G15 = w_active['G15'].value
	G16 = w_active['G16'].value
	G17 = w_active['G17'].value
	G18 = w_active['G18'].value

	e_active['G9'].value = G9
	e_active['G10'].value = G10
	e_active['G11'].value = G11
	e_active['G12'].value = G12
	e_active['G13'].value = G13
	e_active['G14'].value = G14
	e_active['G15'].value = G15
	e_active['G16'].value = G16
	e_active['G17'].value = G17
	e_active['G18'].value = G18


	west_wb.save(west_aug_path)
	east_wb.save(east_aug_path)

