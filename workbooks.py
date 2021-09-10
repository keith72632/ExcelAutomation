import sys
from openpyxl import Workbook, load_workbook
from colors import Prompts

class Books:
	def __init__(self):
		pass
	EXCEPTIONS = 0

	def load_workbooks(self, west_file, east_file, chem_file, table_file, meter_file):
		p = Prompts()
		COLOR = p.PROD if sys.argv[2] == "PROD" else p.DEV
		RESET = p.RESET
		FAIL = p.FAIL
		try:
			west_wb = load_workbook(west_file)
			print(f'{p.ok()}West Workbook loaded from:\n\t{COLOR}{west_file}{RESET}\n')
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}West Workbook loading failed from:\n\t{FAIL}{west_file}{RESET}\n')

		try:
			east_wb = load_workbook(east_file)
			print(f'{p.ok()}East Workbook loaded from:\n\t{COLOR}{east_file}{RESET}\n')
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}East Workbook loading failed from:\n\t{FAIL}{east_file}{RESET}\n')

		try:
			chem_wb = load_workbook(chem_file)
			print(f'{p.ok()}Chemical Treatment Workbook loaded from:\n\t{COLOR}{chem_file}{RESET}\n')
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Chemical Treatment Workbook loading failed from:\n\t{FAIL}{chem_file}{RESET}\n')

		try:
			table_wb = load_workbook(table_file)
			print(f'{p.ok()}Chlorine Table Workbook loaded from:\n\t{COLOR}{table_file}{RESET}\n')
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Chlorine Table Workbook loading failed from:\n\t{FAIL}{table_file}{RESET}\n')

		try:
			meter_wb = load_workbook(meter_file)
			print(f'{p.ok()}Meter Workbook loaded from:\n\t{COLOR}{meter_file}{RESET}\n')
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Meter Workbook loading failed from:\n\t{FAIL}{meter_file}{RESET}\n')

		return west_wb, east_wb, chem_wb, table_wb, meter_wb


	def activate_workbooks(self, west_wb, east_wb, chem_wb):
		return west_wb.active, east_wb.active, chem_wb.active


	def get_sub_workbooks(self, west_wb, chem_wb):
		return west_wb['BMR'], chem_wb['West'], chem_wb['East']


	def save_workbooks(self, west_wb, east_wb, west_path, east_path, chem_wb, chem_path, table_wb, table_path, meter_wb, meter_path):
		p = Prompts()

		try:
			west_wb.save(west_path)
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}West Workbook could not save\n')
		try:
			east_wb.save(east_path)
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}East Workbook could not save\n')

		try:
			chem_wb.save(chem_path)
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Chemical Treatment Record could not save\n')

		try:
			table_wb.save(table_path)
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Chlorine Table could not save\n')

		try:
			meter_wb.save(meter_path)
		except:
			self.EXCEPTIONS += 1
			print(f'{p.err()}Meter Workbook could not save\n')

		print(p.ok() + 'workbooks saved\n')


	def get_exceptions(self):
		return self.EXCEPTIONS