import sys
from openpyxl import Workbook, load_workbook
from lib.colors import Prompts
import traceback
from lib.colors import Prompts
class Books:
	def __init__(self, west_file=None, east_file=None, chem_file=None, table_file=None, meter_file=None):
		p = Prompts()
		self.west_file = west_file
		self.east_file = east_file
		self.chem_file = chem_file
		self.table_file = table_file
		self.meter_file = meter_file
		print('Books constructed')

	def __del__(self):
		print('Books Destroyed')
		self.save_workbooks()

	EXCEPTIONS = 0

	def load_workbooks(self):
		p = Prompts()
		COLOR = p.DEV
		RESET = p.RESET
		FAIL = p.FAIL
		if self.west_file:
			try:
				self.west_wb = load_workbook(self.west_file)
				print(f'{p.ok()}West Workbook loaded from:\n{COLOR}{self.west_file}{RESET}\n')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}West Workbook loading failed from:\n{FAIL}{self.west_file}{RESET}\n')

		if self.east_file:
			try:
				self.east_wb = load_workbook(self.east_file)
				print(f'{p.ok()}East Workbook loaded from:\n{COLOR}{self.east_file}{RESET}\n')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}East Workbook loading failed from:\n{FAIL}{self.east_file}{RESET}\n')

		if self.chem_file:
			try:
				self.chem_wb = load_workbook(self.chem_file)
				print(f'{p.ok()}Chemical Treatment Workbook loaded from:\n{COLOR}{self.chem_file}{RESET}\n')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Chemical Treatment Workbook loading failed from:\n{FAIL}{self.chem_file}{RESET}\n')

		if self.table_file:
			try:
				self.table_wb = load_workbook(self.table_file)
				print(f'{p.ok()}Chlorine Table Workbook loaded from:\n{COLOR}{self.table_file}{RESET}\n')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Chlorine Table Workbook loading failed from:\n{FAIL}{self.table_file}{RESET}\n')

		if self.meter_file:
			try:
				self.meter_wb = load_workbook(self.meter_file)
				print(f'{p.ok()}Meter Workbook loaded from:\n{COLOR}{self.meter_file}{RESET}\n')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Meter Workbook loading failed from:\n{FAIL}{self.meter_file}{RESET}\n')

		return self.west_wb, self.east_wb, self.chem_wb, self.table_wb, self.meter_wb

	def individual_pages_init(self):
		self.bmr_wb = self.west_wb['BMR']
		self.w_chem = self.chem_wb['West']
		self.e_chem = self.chem_wb['East']
		self.west_swor_front = self.west_wb['SWO&R Front Side']
		self.west_swor_back = self.west_wb['SWO&R Back Side']
		self.east_swor_front = self.east_wb['SWO&R Front Side']
		self.east_swor_back = self.east_wb['SWO&R Back Side']
		self.west_ifmr = self.west_wb['IFMR >10,000']
		self.east_ifmr = self.east_wb['IFMR >10,000']
		self.east_table = self.table_wb['East']
		self.west_table = self.table_wb['West']
		self.midnight_readings = self.meter_wb['Midnight']

	def activate_workbooks(self, west_wb, east_wb, chem_wb):
		return west_wb.active, east_wb.active, chem_wb.active


	def get_sub_workbooks(self, west_wb, chem_wb):
		return west_wb['BMR'], chem_wb['West'], chem_wb['East']


	def save_workbooks(self):
		p = Prompts()

		if self.west_file:
			try:
				self.west_wb.save(self.west_file)
				print(f'{p.note()} west workbook saved')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}West Workbook could not save\n')
				# traceback.print_exc()
		
		if self.east_file:
			try:
				self.east_wb.save(self.east_file)
				print(f'{p.note()} east workbook saved')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}East Workbook could not save\n')
				traceback.print_exc()

		if self.chem_file:
			try:
				self.chem_wb.save(self.chem_file)
				print(f'{p.note()} chemical treatment workbook saved')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Chemical Treatment Record could not save\n')
				traceback.print_exc()

		if self.table_file:
			try:
				self.table_wb.save(self.table_file)
				print(f'{p.note()} chlorine table workbook saved')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Chlorine Table could not save\n')
				traceback.print_exc()

		if self.meter_file:
			try:
				self.meter_wb.save(self.meter_file)
				print(f'{p.note()} midnight meter workbook saved')
			except:
				self.EXCEPTIONS += 1
				print(f'{p.err()}Meter Workbook could not save\n')
				traceback.print_exc()

		print(p.ok() + 'workbooks saved')


	def get_exceptions(self):
		return self.EXCEPTIONS